import pandas as pd
import openpyxl
import os
from openpyxl import Workbook
from copy import copy
from openpyxl import load_workbook
import math


wb = openpyxl.load_workbook("jpeg_results_proj.xlsx")
jpeg_sheet = wb["jpeg_data"]
j2k_sheet = wb["j2k_data"]

jpeg_sizes = [cell for row in jpeg_sheet.iter_rows(min_row=2,max_row = 51, min_col=8, max_col=8, values_only=True) for cell in row]
print(jpeg_sizes)
j2k_sizes = [cell for row in j2k_sheet.iter_rows(min_row=2,max_row = 51, min_col=8, max_col=8, values_only=True) for cell in row]
print(j2k_sizes)


# Read the contents of host.cpp
with open("host.cpp", "r") as f:
    host_content = f.readlines()

"""
# Iterate through each line and replace the placeholder with the JPEG size
for i, size in enumerate(jpeg_sizes):
    new_content = []
    for line in host_content:
        if "//INSERT DATASIZE HERE" in line:
            line = f"\tunsigned int dataSize = {size};\n"
        new_content.append(line)

    # Write the modified content to a new file
    new_filename = f"host_jpeg_{i}.cpp"
    with open(new_filename, "w") as f:
        f.writelines(new_content)
    
    print(f"File {new_filename} created with dataSize = {size}")

# Iterate through each line and replace the placeholder with the J2K size
for i, size in enumerate(j2k_sizes):
    new_content = []
    for line in host_content:
        if "//INSERT DATASIZE HERE" in line:
            line = f"\tunsigned int dataSize = {size};\n"
        new_content.append(line)

    # Write the modified content to a new file
    new_filename = f"host_j2k_{i}.cpp"
    with open(new_filename, "w") as f:
        f.writelines(new_content)
    
    print(f"File {new_filename} created with dataSize = {size}")
"""

for i in range(900,1601):
    new_content = []
    size = i 
    for line in host_content:
        if "//INSERT DATASIZE HERE" in line:
            line = f"\tunsigned int dataSize = {size};\n"
        new_content.append(line)

    # Write the modified content to a new file
    new_filename = f"host_range_{i}.cpp"
    with open(new_filename, "w") as f:
        f.writelines(new_content)
    
    print(f"File {new_filename} created with dataSize = {size}")
